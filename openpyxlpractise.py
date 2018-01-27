from openpyxl import Workbook
from datetime import *
from openpyxl.styles import numbers
wb=Workbook()

dest_filename='excelpra.xlsx'

ws1=wb.active

for row in range(1,10):
    
    ws1.cell(column=1,row=row,value=date(2010,7,21))
    ws1.cell(column=1,row=row).number_format = numbers.FORMAT_DATE_XLSX14
    numbers.
wb.save(filename=dest_filename)
